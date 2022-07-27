#import time, sys, unittest, random, json, requests, openpyxl, testlink
import time, json, random, platform,openpyxl
from datetime import datetime
from appium import webdriver
from random import randint
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert
import time
import inspect
from appium.webdriver.common.touch_action import TouchAction
from sys import exit
from openpyxl import Workbook




class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"





#now = datetime.now()
#date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
#date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]

if platform == "linux" or platform == "linux2":
    local = "/home/oem/groupware-auto-test"
    json_file = local + "/appium_mail_app.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "/Log/"
    execution_log = local + log_folder + "mailapp_execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("mailapp_execution_log_", "fail_log_")
    error_log = execution_log.replace("mailapp_execution_log_", "error_log_") 
else :
    local = "D:\\File_Du_Lieu\\Automation Test\\LuuNgo_Appium"
    json_file = local + "\\appium_mail_app.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "\\Log\\"
    execution_log = local + log_folder + "mailapp_execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("mailapp_execution_log_", "fail_log_")
    error_log = execution_log.replace("mailapp_execution_log_", "error_log_")

testcase_log = local + log_folder + "testcase_mail_app_" + str(objects.date_id) + ".xlsx"   

logs = [execution_log,fail_log,error_log,testcase_log]
for log in logs: 
    if".txt" in log:
        open(log,"x").close()
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        wb.save(log)




# Connect to Appium with the below desire capabilities
# http://appium.io/docs/en/writing-running-appium/caps/
dc = {
    "deviceName": "ad091603510be6a1a9",
    "platformName": "Android",
    "app": local + "\\mail-app-hanbiro-release.apk",
    "automationName": "UiAutomator2",
    "autoGrantPermissions": "true",
    "appWaitPackage": "com.hanbiro.mailapp",
    "appWaitActivity": "com.hanbiro.mailapp.MainActivity"    

}

# If desire capabilities are valid, the app will be open at Log in screen
driver = webdriver.Remote('http://localhost:4723/wd/hub', dc)
now = datetime.now()
mail_title = "Mail App is write at"  + str(now)

# Input information for log-in

#with open("D:\\File_Du_Lieu\\Selenium\\Selenium_python\\Selenium_python\\Appium-pyhton\\LuuNgo_Appium\\appium_mail_app.json") as json_data_file:
#    data = json.load(json_data_file)

def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def ValidateFailResultAndSystem(fail_msg):
    Logging(fail_msg)
    append_fail_result = open(fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()
    


def TestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging("description")
    if status=="Pass":
        print(objects.testcase_pass)
    else:
        print(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows))+1
    current_sheet.cell(row=start_row,column=1).value=menu
    current_sheet.cell(row=start_row,column=2).value=sub_menu
    current_sheet.cell(row=start_row,column=3).value=testcase
    current_sheet.cell(row=start_row,column=4).value=status
    current_sheet.cell(row=start_row,column=5).value=description
    current_sheet.cell(row=start_row,column=6).value=objects.date_time
    current_sheet.cell(row=start_row,column=7).value= tester
    
    wb.save(testcase_log)



def log_in_mail_app():

    
    Logging("----------------1. Log In Mail App-----------------------------")
    #check_crash= WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, data["domain_input"])))
    try:
        check_crash= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["domain_input"])))
        if  check_crash.is_displayed():
            
            Logging("=>=> No Crash App") 
        else:
            Logging("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        txt_domain= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["domain_input"])))
        if  txt_domain.is_displayed():
            Logging("=>=> No Crash App")
            txt_domain.send_keys(data["login_page"])
            Logging("1.Input Domain") 
        else:
            ValidateFailResultAndSystem("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)

    try:
        username = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["txt_id"])))
        if  username.is_displayed():
            Logging("=>=> No Crash App") 
            username.send_keys(data["input_user_user"])
            Logging("2.Input ID")
        else:
            Logging("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    try:
        password = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["txt_pw"])))
        if  password.is_displayed():
            Logging("=>=> No Crash App") 
            password.send_keys(data["input_pw"])
            Logging("3.Input Password")
        else:
            Logging("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    time.sleep(1)
    button_log_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["btn_log_in"])))
    button_log_in.click()
    Logging("4.Click Log In button")
    time.sleep(5)
    try:
        click_btn_create_new = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_write_mail"])))
        if  click_btn_create_new.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["mail_app"]["log_in_mail_app"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["mail_app"]["log_in_mail_app"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    if "Inbox" or "받은메일함" or "Hộp thứ đến" in driver.page_source :
        Logging("2. Log in successfully")
    else:
        Logging("2. Crash app")
        exit(0)
    time.sleep(1)

    if 'Inbox' in driver.page_source :
        Logging("2. English language is used")
    else:
        Logging("2. language Orther")

        try:
            click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
            if  click_icon_list_mail.is_displayed():
                Logging("=> No Crash App") 
                click_icon_list_mail.click()
                Logging("1. Click User Setting => Pass")
            else:
                ValidateFailResultAndSystem("=> Crash App")
                exit(0)
        except WebDriverException:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
        time.sleep(1)
        click_avatar_new = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["avatar_user"])))
        click_avatar_new.click()
        Logging("2. Click Avatar => Pass")
        click_langue_list = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["click_change_langue"])))
        click_langue_list.click()

        try:
            click_list_language = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_langue_eng"])))
            if  click_list_language.is_displayed():
                Logging("=> No Crash App") 
                click_list_language.click()
                Logging("3. Change langue => Pass")
            else:
                ValidateFailResultAndSystem("=> Crash App")
                exit(0)
        except WebDriverException:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
        

def send_mail_app():
    Logging("------------------------------------------------------Send Mail------------------------------------------------------")
    try:
        click_btn_create_new = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_write_mail"])))
        if  click_btn_create_new.is_displayed():
            Logging("=> No Crash App") 
            click_btn_create_new.click()
            time.sleep(1)
            Logging("1. Click button Create new  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    input_recipient_email = driver.find_element_by_class_name("android.widget.EditText")
    input_recipient_email.send_keys("automationtest1@qa.hanbiro.net" +  "   ")
    Logging("2. Input recipient email  successfully")
    if 'automationtest1@qa.hanbiro.net' in driver.page_source :
        Logging("2.TO =>  PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["input_recipient_mail"]["pass"])
    else:
        Logging("2.TO =>  FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["input_recipient_mail"]["fail"])
    time.sleep(1)
    Logging("------------------------------------------------------CC------------------------------------------------------")
    time.sleep(1)
    try:
        icon_org_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["org_bcc_cc"])))
        if  icon_org_cc.is_displayed():
            Logging("=> No Crash App") 
            icon_org_cc.click()
            time.sleep(1)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(4)
    try:
        select_tab_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["tab_cc"])))
        if  select_tab_cc.is_displayed():
            Logging("=> No Crash App") 
            select_tab_cc.click()
            time.sleep(3)
            txt_search_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_search"])))
            txt_search_contact_org.click()
            time.sleep(3)
            driver.is_keyboard_shown()
            time.sleep(2)
            #driver.press_keycode(29)
            #time.sleep(2)
            #driver.press_keycode(49)
            #time.sleep(2)
            driver.press_keycode(48)
            time.sleep(3)
            driver.press_keycode(43)
            time.sleep(3)
            driver.press_keycode(41)
            time.sleep(2)
            driver.press_keycode(29)
            time.sleep(2)
            driver.press_keycode(48)
            time.sleep(2)
            driver.press_keycode(37)
            time.sleep(2)
            driver.press_keycode(43)
            time.sleep(2)
            driver.press_keycode(42)
            time.sleep(2)
            #driver.press_keycode(48)
            #driver.press_keycode(33)
            #driver.press_keycode(47)
            #driver.press_keycode(48)
            #driver.press_keycode(8)
            driver.press_keycode(66)
            time.sleep(3)
            Logging("3. Search User  successfully")
            select_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_user_cc_auto1"])))
            select_contact_org.click()
            time.sleep(1)
            Logging("4. Click User in Tab Cc successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    #txt_search_user = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_search"])))
    #txt_search_user.send_keys(data["send_mail"]["user_search"])
    #txt_search_user.send_keys(Keys.ENTER)
    time.sleep(2)
    icon_check_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_confirm_select_user"])))
    icon_check_confirm.click()
    time.sleep(1)
    Logging("7. Check icon successfully")
    try:
        icon_show_view_mail_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["send_mail"]["icon_view_mail_cc"])))
        if  icon_show_view_mail_cc.is_displayed():
            Logging("=> No Crash App") 
            icon_show_view_mail_cc.click()
            time.sleep(1)
            Logging("8. Check icon show mail CC successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    #icon_show_view_mail_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["send_mail"]["icon_view_mail_cc"])))
    #icon_show_view_mail_cc.click()
    #Logging("8. Check icon show mail CC successfully")
    time.sleep(3)
    if 'AutomationTest' in driver.page_source :
        Logging("Add User Cc =>  PASS")
    else:
        Logging("Add User Cc =>  FAIL")
    time.sleep(1)
    Logging("------------------------------------------------------Tab BCC------------------------------------------------------")
    icon_org_bcc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["org_bcc_cc"])))
    icon_org_bcc.click()
    Logging("3. Click Organization  successfully")
    #select_tab_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["tab_bcc"])))
    #select_tab_cc.click()
    #Logging("4. Click Tab Cc successfully")
    time.sleep(2)
    try:
        select_tab_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["tab_bcc"])))
        if  select_tab_cc.is_displayed():
            Logging("=> No Crash App") 
            select_tab_cc.click()
            time.sleep(2)
            txt_search_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_search"])))
            txt_search_contact_org.click()
            time.sleep(2)
            driver.is_keyboard_shown()
            time.sleep(2)
            #driver.press_keycode(29)
            #time.sleep(2)
            #driver.press_keycode(49)
            #time.sleep(2)
            driver.press_keycode(48)
            time.sleep(3)
            driver.press_keycode(43)
            time.sleep(2)
            driver.press_keycode(41)
            time.sleep(2)
            driver.press_keycode(29)
            time.sleep(2)
            driver.press_keycode(48)
            time.sleep(2)
            driver.press_keycode(37)
            time.sleep(1)
            #driver.press_keycode(37)
            #driver.press_keycode(43)
            #driver.press_keycode(41)
            #driver.press_keycode(48)
            #driver.press_keycode(33)
            #driver.press_keycode(47)
            #driver.press_keycode(48)
            #driver.press_keycode(8)
            driver.press_keycode(66)
            time.sleep(5)
            Logging("3. Search User  successfully")
            select_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_user_cc_auto2"])))
            select_contact_org.click()
            time.sleep(2)
            Logging("4. Click User in Tab Cc successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(2)
    icon_check_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_confirm_select_user"])))
    icon_check_confirm.click()
    Logging("7. Check icon successfully")
    time.sleep(2)
    if 'AutomationTest2' in driver.page_source :
        Logging("2.BCc =>  PASS")
    else:
        Logging("2.BCC =>  FAIL")

    time.sleep(2)
    subject_email = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, data["send_mail"]["txt_title_mail_app"])))
    subject_email.send_keys(mail_title)
    Logging("9. Input Title Mail App successfully")
    time.sleep(1)
    Logging("------------------------------------------------------Input Content------------------------------------------------------")
    #editor_frame = driver.find_element_by_class_name("android.widget.RelativeLayout")
    txt_content = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_xpath_content"])))
    txt_content.send_keys(data["send_mail"]["content_mail"])
    Logging("10. Input Content successfully")
    time.sleep(2)
    Logging("------------------------------------------------------Attach File------------------------------------------------------")
    click_icon_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_attach_file"])))
    click_icon_attach_file.click()
    time.sleep(2)
    Logging("10. Check icon attach file successfully")
    time.sleep(1)
    try:
        click_folder_choose_photo = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["folder_choose_photo"])))
        if  click_folder_choose_photo.is_displayed():
            Logging("=> No Crash App") 
            click_folder_choose_photo.click()
            time.sleep(2)
            Logging("11. Check folder Choose Photo successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        click_folder_download = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["folder_dowload"])))
        if  click_folder_download.is_displayed():
            Logging("=> No Crash App") 
            click_folder_download.click()
            time.sleep(2)
            Logging("12. Check folder Screenshotsuccessfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    try:
        select_file_attach = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["send_mail"]["select_image_mail_app"])))
        if  select_file_attach.is_displayed():
            Logging("=> No Crash App") 
            select_file_attach.click()
            time.sleep(1)
            select_file2_attach = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_image2_mail_app"])))
            select_file2_attach.click()
            time.sleep(2)
            Logging("13. Check file successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(2)
    
    try:
        click_icon_confirm_attachfile= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["send_mail"]["icon_confirm_attach_file"])))
        if  click_icon_confirm_attachfile.is_displayed():
            Logging("=> No Crash App") 
            click_icon_confirm_attachfile.click()
            time.sleep(2)
            Logging("14. Click Icon Confirm attach file  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    #click_folder_choose_photo = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["folder_choose_photo"])))
    #click_folder_choose_photo.click()
    #time.sleep(1)
    #Logging("11. Check folder Choose Photo successfully")
    #click_folder_screeshots = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["folder_screenshots"])))
    #click_folder_screeshots.click()
    #time.sleep(2)
    #Logging("12. Check folder Screenshotsuccessfully")
    #select_file_attach = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, data["send_mail"]["select_image_mail_app"])))
    #select_file_attach.click()
    #time.sleep(1)
    #Logging("13. Check file successfully")
    #click_icon_confirm_attachfile= WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, data["send_mail"]["icon_confirm_attach_file"])))
    #click_icon_confirm_attachfile.click()
    #Logging("14. Click Icon Confirm attach file  successfully")
    #time.sleep(1)


    
    driver.swipe(start_x=523, start_y=1778, end_x=523, end_y=1089, duration=800)
    time.sleep(3)
    Logging("=>Scroll successfully")
    time.sleep(1)
    title_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["send_mail"]["image_attache_comfirm"])))
    time.sleep(1)
    if title_attach_file.is_displayed():
        Logging("=> Attach File successfully")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["attach_file_mail"]["pass"])
    else:
        Logging("=> Attach File Fail")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["attach_file_mail"]["fail"])
    icon_send_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_send_mail_app"])))
    icon_send_mail.click()
    time.sleep(2)
    Logging("15. Check icon Send mail  successfully")
    time.sleep(20)
   
    #driver.swipe(start_x=0, start_y = 0, end_x = 0, end_y = 1000, duration=800)


    
    
    try:
        driver.swipe(start_x=0, start_y = 0, end_x = 0, end_y = 1000, duration=800)
        push_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["check_push_mail_app"])))
        push_mail_app.click()
        Logging("Show push")
    except WebDriverException:
        Logging("Not show push")
    time.sleep(2)
    driver.swipe(start_x=0, start_y = 1000, end_x = 0, end_y = 0, duration=800)


    time.sleep(2)
    try:
        click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
        if  click_icon_list_mail.is_displayed():
            Logging("=> No Crash App") 
            click_icon_list_mail.click()
            time.sleep(2)
            Logging("1. Click Icon List Mail  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        click_folder_secret_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_secret"])))
        if  click_folder_secret_mail_app.is_displayed():
            Logging("=> No Crash App") 
            click_folder_secret_mail_app.click()
            time.sleep(2)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(2)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(2)
    Logging("1. Click Inbox successfully")
    time.sleep(2)
    if 'Mail App' in driver.page_source :
        Logging("=> Send Mail => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["send_mail"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Send Mail => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["send_mail"]["fail"])
    time.sleep(2)
    #count_mail = int(len(driver.find_elements_by_xpath(data["send_mail"]["count_mail_inbox"])))
    #Logging("Total mail:",count_mail)
    #time.sleep(3)
    Logging("------------------------------------------------------View Mail details------------------------------------------------------")
    click_title_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_title_mail_view"])))
    click_title_mail.click()
    time.sleep(3)
    Logging("1. Click title mail successfully")
    try:
        click_view_details_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["view_detail_mail"])))
        if  click_view_details_mail.is_displayed():
            Logging("=> No Crash App") 
            click_view_details_mail.click()
            time.sleep(2)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(3)

    if 'From' in driver.page_source :
        Logging("=> View Mail details => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_mail_detail"]["pass"])

    else:
        Logging("=> View Mail details => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_mail_detail"]["fail"])
    time.sleep(3)
    Logging("------------------------------------------------------View Mail content------------------------------------------------------")
    if 'Luu Luu' in driver.page_source :
        Logging("=> View Mail content => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_mail_content"]["pass"])
    else:
        ValidateFailResultAndSystem("=> View Mail content => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_mail_content"]["fail"])
    time.sleep(2)

    Logging("------------------------------------------------------View Attach file------------------------------------------------------")
    click_file_image_view = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["view_image"])))
    click_file_image_view.click()
    time.sleep(5)
    show_image_in_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["show_image"])))
    if show_image_in_mail.is_displayed():
        Logging("==> View the Image successfully")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_attach_file"]["pass"])
    else:
        Logging("==> View the Image Fail")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["view_attach_file"]["fail"])
    time.sleep(2)
    Logging("=> View Attach file successfully")
    time.sleep(2)
    Logging("------------------------------------------------------Download file------------------------------------------------------")
    click_icon_download_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_download_file"])))
    click_icon_download_file.click()
    time.sleep(5)
    if 'Save albums success' in driver.page_source :
        Logging("=> Download file => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["download_file"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Download file => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["download_file"]["fail"])
    time.sleep(1)
    click_icon_ok_download_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_ok_download_file"])))
    click_icon_ok_download_file.click()
    time.sleep(1)
    Logging("------------------------------------------------------Share file from MailApp------------------------------------------------------")

    try:
        click_icon_share_file_from_mailapp = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_share_file_from_mailapp"])))
        if  click_icon_share_file_from_mailapp.is_displayed():
            Logging("=> No Crash App") 
            click_icon_share_file_from_mailapp.click()
            time.sleep(3)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(3)

    #click_icon_share_file_from_mailapp = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_share_file_from_mailapp"])))
    #click_icon_share_file_from_mailapp.click()
    
    if 'Share file from MailApp' in driver.page_source :
        Logging("=> Share file from MailApp => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["share_file"]["pass"])
    else:
        Logging("=> Share file from MailApp => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["share_file"]["fail"])
    
    time.sleep(1)
    driver.back()
    Logging("1. Click back successfully")
    time.sleep(1)
    driver.back()
    #click_icon_back = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_view_image"])))
    #click_icon_back.click()
    time.sleep(2)
    Logging("1. Click Icon Back successfully")
    
    Logging("------------------------------------------------------Forward Mail------------------------------------------------------")
    time.sleep(1)
    click_forward_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["click_icon_forward"])))
    click_forward_mail.click()
    Logging("1. Click Icon Forward Mail successfully")
    time.sleep(3)
    try:
        icon_org_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["org_bcc_cc"])))
        if  icon_org_cc.is_displayed():
            Logging("=> No Crash App") 
            icon_org_cc.click()
            time.sleep(2)
            Logging("2. Click Organization  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(3)
    txt_search_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_search"])))
    txt_search_contact_org.click()
    time.sleep(1)
    driver.is_keyboard_shown()
    driver.press_keycode(48)
    time.sleep(2)
    driver.press_keycode(43)
    time.sleep(3)
    driver.press_keycode(41)
    time.sleep(2)
    driver.press_keycode(29)
    time.sleep(1)
    driver.press_keycode(48)
    time.sleep(1)
    driver.press_keycode(37)
    time.sleep(1)
    driver.press_keycode(66)
    time.sleep(5)
    Logging("3. Search User  successfully")
    select_contact_org = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_user_cc_auto1"])))
    select_contact_org.click()
    time.sleep(1)
    Logging("4. Click User in Tab Cc successfully")
    time.sleep(3)

    #icon_org_cc = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["org_bcc_cc"])))
    #icon_org_cc.click()
    #Logging("2. Click Organization  successfully")

    icon_check_confirm = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_confirm_select_user"])))
    icon_check_confirm.click()
    Logging("5. Check icon successfully")
    time.sleep(2)
    icon_send_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_send_mail_app"])))
    icon_send_mail.click()
    time.sleep(2)
    Logging("6. Check icon Send mail  successfully")
    time.sleep(1)
    icon_back_fw_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_fw"])))
    icon_back_fw_mail.click()
    time.sleep(1)
    Logging("6. Check icon Send mail  successfully")
    time.sleep(3)
    click_icon_list_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    Logging("1. Click Icon List Mail  successfully")
    click_folder_secret_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_secret"])))
    click_folder_secret_mail_app.click()
    time.sleep(2)
    click_icon_list_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    click_folder_inbox_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(2)
    Logging("1. Click Inbox successfully")
    time.sleep(2)
    if 'FW' in driver.page_source :
        Logging("=> Forward Mail => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["forward_mail"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Forward Mail => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["forward_mail"]["fail"])
    time.sleep(2)
    Logging("------------------------------------------------------HMail-7 : Mark important------------------------------------------------------")
    #time.sleep(2)
    click_icon_important = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_important"])))
    click_icon_important.click()
    time.sleep(2)
    Logging("1. Click Icon Mark important  successfully")
    Logging("------------------------------------------------------HMail-8 : Mark as Read ------------------------------------------------------")
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    Logging("1. Click Icon List Mail  successfully")
    time.sleep(1)
    '''
    total=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["number_mark_as_unread"])))
    text_total = total.text
    time.sleep(1)
    Logging(text_total)
    Logging("---  Total before Mark as unread : " + text_total)
    time.sleep(3)
    '''
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(1)
    Logging("1. Click Inbox successfully")

    time.sleep(1)
    click_icon_mail_mask_as_read = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["title_mail_mark_as_unread"])))
    click_icon_mail_mask_as_read.click()
    time.sleep(3)
    Logging("2. Click Icon  successfully")
    #click_btn_mask_as_read = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_mark_as_read"])))
    #click_btn_mask_as_read.click()
    #time.sleep(1)
    #Logging("2. Click button Mark as Read  successfully")
    time.sleep(3)
    try:
        click_btn_mask_as_read = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_mark_as_read"])))

        if  click_btn_mask_as_read.is_displayed():
            Logging("=> No Crash App") 
            click_btn_mask_as_read.click()
            time.sleep(2)
            Logging("3.  Mark as UnRead successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
        time.sleep(1)
    except WebDriverException:
        Logging("3.  Mark as UnRead Fail")
    time.sleep(3)
    click_icon_back = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_mark_read"])))
    click_icon_back.click()
    time.sleep(1)
    Logging("3. Click Icon Back successfully")
    Logging("------------------------------------------------------HMail-9 : Mark as UnRead ------------------------------------------------------")
    time.sleep(1)
    click_icon_mail_mask_as_read = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["title_mail_mark_as_unread"])))
    click_icon_mail_mask_as_read.click()
    time.sleep(3)
    Logging("1. Click Icon  successfully")
    try:
        click_btn_mask_as_read = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_mark_as_read"])))
        click_btn_mask_as_read.click()
        time.sleep(2)
        Logging("2.  Mark as UnRead successfully")
    except WebDriverException:
        Logging("2.  Mark as UnRead Fail")
    time.sleep(3)
    click_icon_back = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_mark_read"])))
    click_icon_back.click()
    time.sleep(2)
    Logging("3. Click Icon Back successfully")
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(1)
    Logging("4. Click Icon List Mail  successfully")
    #total=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["number_mark_as_unread"])))
    #text_total1 = total.text
    #time.sleep(1)
    #Logging(text_total1)
    #Logging("---  Total after Mark as read : " + text_total1)
    time.sleep(3)
    Logging("------------------------------------------------------HMail-16 : Search Mail ------------------------------------------------------")
    time.sleep(2)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(2)
    Logging("1. Click Inbox successfully")
    time.sleep(2)
    click_icon_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_search_mail"])))
    click_icon_search_mail.click()
    Logging("2. Click icon search mail successfully ")  
    time.sleep(2)
    try:
        txt_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["txt_search_mail"])))
        if  txt_search_mail.is_displayed():
            Logging("=> No Crash App") 
            txt_search_mail.send_keys(data["send_mail"]["title_mail_search"])
            Logging("3. Input title mail search successfully ")  
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    time.sleep(2)
    #txt_search_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["txt_search_mail"])))
    #txt_search_mail.send_keys(data["send_mail"]["title_mail_search"])
    #Logging("3. Input title mail search successfully ")  

    time.sleep(2)
    if 'Mail App' in driver.page_source :
        Logging("=> Search Mail => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["search_mail"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Search Mail => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["search_mail"]["fail"])

    time.sleep(2)
    Logging("------------------------------------------------------Reply All Mail------------------------------------------------------")
    time.sleep(1)
    Logging("1. Click Inbox Mail  successfully")
    click_title_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail_reply"])))
    click_title_mail.click()
    time.sleep(2)
    Logging("1. Select Mail Reply All successfully")
    click_reply_all_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["click_icon_reply_all"])))
    click_reply_all_mail.click()
    Logging("2. Click Icon Reply  successfully")
    time.sleep(1)
    try:
        content_email_reply_all = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_content_mail_reply_all"])))
        if  content_email_reply_all.is_displayed():
            Logging("=> No Crash App") 
            content_email_reply_all.send_keys(data["send_mail"]["content_mail_reply_all"])
            Logging("3. Input Mail Reply successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    #content_email_reply_all = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_content_mail_reply_all"])))
    #content_email_reply_all.send_keys(data["send_mail"]["content_mail_reply_all"])
    #Logging("3. Input Mail Reply successfully")
    time.sleep(2)
    icon_send_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_send_mail_app"])))
    icon_send_mail.click()
    time.sleep(2)
    icon_back_fw_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_fw"])))
    icon_back_fw_mail.click()
    time.sleep(3)
    Logging("4. Click icon Back successfully ")  
    click_icon_back_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_search_mail"])))
    click_icon_back_search_mail.click()
    Logging("5. Click icon Back Search successfully ")  
    time.sleep(3)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    Logging("1. Click Icon List Mail  successfully")
    click_folder_secret_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_secret"])))
    click_folder_secret_mail_app.click()
    time.sleep(2)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(1)
    Logging("1. Click Inbox successfully")
    time.sleep(2)
    if 'RE' in driver.page_source :
        Logging("=> REPLY ALL Mail => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["reply_mail"]["pass"])
    else:
        ValidateFailResultAndSystem("=> REPLY ALL Mail => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["reply_mail"]["fail"])
    time.sleep(1)
    Logging("------------------------------------------------------Report Spam------------------------------------------------------")
    time.sleep(1)
    # Report Spam
    time.sleep(2)
    select_icon_mail_spam = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["title_mail_mark_as_unread"])))
    select_icon_mail_spam.click()
    time.sleep(3)
    Logging("1. Select email  successfully")
    click_icon_list_spam_move= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_spam_move_cancel"])))
    click_icon_list_spam_move.click()
    Logging("2. Click Icon List spam/Move successfully")
    time.sleep(3)
    try:
        select_spam_mail= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_report_spam"])))
        if  select_spam_mail.is_displayed():
            Logging("=> No Crash App") 
            select_spam_mail.click()
            time.sleep(2)
            Logging("3. Select Spam mail successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    #select_spam_mail= WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_report_spam"])))
    #select_spam_mail.click()
    #time.sleep(1)
    #Logging("3. Select Spam mail successfully")
    time.sleep(3)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(4)
    Logging("4. Click Icon List Mail  successfully")
    click_folder_spam_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_spam"])))
    click_folder_spam_mail_app.click()
    time.sleep(2)
    if 'Mail App' in driver.page_source :
        Logging("=> Report Spam => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["report_mail"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Report Spam => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["report_mail"]["fail"])
    time.sleep(1)

    Logging("------------------------------------------------------Move Mail------------------------------------------------------")
    # Delete Mail
    time.sleep(1)
    select_icon_mail_spam = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["title_mail_mark_as_unread"])))
    select_icon_mail_spam.click()
    time.sleep(1)
    Logging("1. Select email  successfully")
    #click_icon_delete_mail= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_icon_delete"])))
    #click_icon_delete_mail.click()
    #Logging("2. Click Icon Delete mail successfully")
    time.sleep(1)
    click_icon_list_spam_move= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_spam_move_cancel"])))
    click_icon_list_spam_move.click()
    Logging("2. Click Icon List spam/Move successfully")
    time.sleep(3)
    try:
        select_move_mail= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail_move"])))
        if  select_move_mail.is_displayed():
            Logging("=> No Crash App") 
            select_move_mail.click()
            time.sleep(2)
            Logging("3. Click Value Move successfully")
            move_mail_draft= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_draft"])))
            move_mail_draft.click()
            time.sleep(2)
            Logging("4. Select folder Drafts successfully")
            click_btn_done_move= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_done_move_draft"])))
            click_btn_done_move.click()
            time.sleep(2)
            Logging("5. Click button Done successfully")
            time.sleep(2)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(3)
    if 'Mail App' in driver.page_source :
        Logging("=> Move Mail => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["move_mail"]["fail"])
    else:
        ValidateFailResultAndSystem("=> Move Mail => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["move_mail"]["pass"])
    time.sleep(3)
    Logging("------------------------------------------------------Reply Mail------------------------------------------------------")
    time.sleep(1)
    # Reply Mail
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(3)
    Logging("1. Click Inbox successfully")
    time.sleep(3)
    click_icon_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_search_mail"])))
    click_icon_search_mail.click()
    time.sleep(3)
    Logging("2. Click icon search mail successfully ")    
    txt_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["txt_search_mail"])))
    txt_search_mail.send_keys(data["send_mail"]["title_mail_search"])
    Logging("3. Input title mail search successfully ")  
    time.sleep(3)
    click_title_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail_reply"])))
    click_title_mail.click()
    time.sleep(2)
    Logging("1. Select Mail Reply  successfully")
    click_reply_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["click_icon_reply"])))
    click_reply_mail.click()
    Logging("2. Click Icon Reply  successfully")
    time.sleep(1)
    try:
        content_email_reply = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_content_mail_reply"])))
        if  content_email_reply.is_displayed():
            Logging("=> No Crash App") 
            content_email_reply.send_keys(data["send_mail"]["content_mail_reply"])
            Logging("3. Input Mail Reply successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    #content_email_reply = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_content_mail_reply"])))
    #content_email_reply.send_keys(data["send_mail"]["content_mail_reply"])
    #Logging("3. Input Mail Reply successfully")

    time.sleep(3)
    icon_send_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_send_mail_app"])))
    icon_send_mail.click()
    time.sleep(2)
    Logging("15. Check icon Send mail  successfully")
    time.sleep(3)
    icon_back_fw_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_fw"])))
    icon_back_fw_mail.click()
    time.sleep(3)
    click_icon_back_search_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_back_search_mail"])))
    click_icon_back_search_mail.click()
    Logging("5. Click icon Back Search successfully ")  
    time.sleep(3)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(1)
    Logging("1. Click Icon List Mail  successfully")
    click_folder_secret_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_secret"])))
    click_folder_secret_mail_app.click()
    time.sleep(3)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(3)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_inbox"])))
    click_folder_inbox_mail_app.click()
    time.sleep(3)
    if 'RE' in driver.page_source :
        Logging("=> REPLY Mail => PASS")
    else:
        ValidateFailResultAndSystem("=> REPLY Mail => FAIL")
    time.sleep(3)

    Logging("------------------------------------------------------View Mail List in Inbox------------------------------------------------------")
    if 'Mail App' in driver.page_source :
        Logging("=> View Mail content => PASS")
    else:
        Logging("=> View Mail content => FAIL")
    time.sleep(2)
    Logging("------------------------------------------------------View Mail List in folder Sent Mail------------------------------------------------------")
    time.sleep(1)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    Logging("4. Click Icon List Mail  successfully")
    click_folder_sent_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_report_sent_mail"])))
    click_folder_sent_mail_app.click()
    time.sleep(3)
    if 'Mail App' in driver.page_source :
        Logging("=> View Mail List in folder Sent Mail => PASS")
    else:
        Logging("=> View Mail List in folder Sent Mail => FAIL")
    time.sleep(2)
    

    Logging("------------------------------------------------------Delete multiple email------------------------------------------------------")
    time.sleep(1)
    select_mail_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["title_mail_mark_as_unread"])))
    select_mail_delete.click()
    time.sleep(2)
    select_mail2_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["mail2_delete"])))
    select_mail2_delete.click()
    time.sleep(2)
    select_mail3_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["mail3_delete"])))
    select_mail3_delete.click()
    time.sleep(2)
    select_mail4_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["mail4_delete"])))
    select_mail4_delete.click()
    time.sleep(2)
    Logging("1. Select multiple email  successfully")
    click_icon_delete_mail= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_icon_delete"])))
    click_icon_delete_mail.click()
    time.sleep(2)
    Logging("2. Click Icon Delete mail successfully")
    time.sleep(3)
    

def vacation_auto_replies_mail_app():

    Logging("------------------------------------------------------Vacation Auto Replies------------------------------------------------------")
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    click_icon_list_mail.click()
    time.sleep(2)
    driver.swipe(start_x=523, start_y=1778, end_x=523, end_y=1089, duration=800)
    Logging("1. Scroll to Settings successfully")
    time.sleep(2)
    click_folder_setting_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_setting"])))
    click_folder_setting_mail_app.click()
    Logging("2. Click Settings successfully")
    time.sleep(2)

    '''
    try:
        click_vacation_auto_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_vacation_auto_mail_app"])))
        if  click_vacation_auto_mail_app.is_displayed():
            Logging("=> No Crash App") 
            click_vacation_auto_mail_app.click()
            Logging("3. Select Vacation Auto Replies successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    '''
    click_vacation_auto_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_vacation_auto_mail_app"])))
    click_vacation_auto_mail_app.click()
    #time.sleep(1)
    #Logging("3. Select Vacation Auto Replies successfully")
    time.sleep(3)
    turn_on_auto_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["turn_on_vacation_auto_mail_app"])))
    turn_on_auto_mail_app.click()
    Logging("4. Turn On Vacation Auto Replies successfully")
    time.sleep(3)
    try:
        click_icon_calendar_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_icon_calendar"])))
        if  click_icon_calendar_mail_app.is_displayed():
            Logging("=> No Crash App") 
            click_icon_calendar_mail_app.click()
            Logging("5. Click Icon Calendar successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    time.sleep(3)
    #click_icon_calendar_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_icon_calendar"])))
    #click_icon_calendar_mail_app.click()
    try:
        click_icon_next_month_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_next_month_calendar"])))
        if  click_icon_next_month_mail_app.is_displayed():
            Logging("=> No Crash App") 
            click_icon_next_month_mail_app.click()
            time.sleep(2)
            Logging("6. Click Icon Next Month successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    time.sleep(2)
    #click_icon_next_month_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_next_month_calendar"])))
    #click_icon_next_month_mail_app.click()
    #Logging("6. Click Icon Next Month successfully")
    select_date_vacation = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_date_vacation_auto_mail_app"])))
    select_date_vacation.click()
    time.sleep(2) 
    Logging("7. Select Date successfully")
    click_btn_ok_date_vacation = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_date_vacation_btn_ok"])))
    click_btn_ok_date_vacation.click()
    time.sleep(2) 
    Logging("8. Click button OK successfully")
    time.sleep(2) 
    try:
        txt_input_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_message"])))
        if  txt_input_message.is_displayed():
            Logging("=> No Crash App") 
            txt_input_message.send_keys(data["send_mail"]["content_message"])
            Logging("9. Input Mesage successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    #txt_input_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_message"])))
    #txt_input_message.send_keys(data["send_mail"]["content_message"])
    #Logging("9. Input Mesage successfully")
    time.sleep(3) 
    click_btn_save_vacation_auto = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_save_vacation_auto_replies"])))
    click_btn_save_vacation_auto.click()
    time.sleep(2) 
    Logging("10. Click button SAVE successfully")
    time.sleep(1) 
    click_vacation_auto_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_vacation_auto_mail_app"])))
    click_vacation_auto_mail_app.click()
    time.sleep(2)
    Logging("11. Select Vacation Auto Replies successfully")
    time.sleep(2)
    if 'Please wait' in driver.page_source :
        Logging("=> Vacation Auto Replies => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["vacation_auto_replies"]["pass"])
    else:
        Logging("=> Vacation Auto Replies => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["vacation_auto_replies"]["fail"])
    time.sleep(1)

    time.sleep(1) 
    turn_off_auto_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, data["send_mail"]["turn_off_vacation_auto_mail_app"])))
    turn_off_auto_mail_app.click()
    time.sleep(2)
    Logging("12. Turn OFF Vacation Auto Replies successfully")
    time.sleep(2)
    click_btn_save_vacation_auto = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_save_vacation_auto_replies"])))
    click_btn_save_vacation_auto.click()
    time.sleep(2) 
    Logging("13. Click button SAVE successfully")
    time.sleep(2) 
    
def auto_sort_mail_app():
    
    Logging("------------------------------------------------------Auto-Sort------------------------------------------------------")
    #click_icon_list_mail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))
    #click_icon_list_mail.click()
    #time.sleep(2)
    #driver.swipe(start_x=523, start_y=1778, end_x=523, end_y=1089, duration=800)
    #Logging("1. Scroll to Settings successfully")
    #time.sleep(3)
    #click_folder_setting_mail_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_setting"])))
    #click_folder_setting_mail_app.click()
    time.sleep(2)
    Logging("2. Click Settings successfully")
    click_auto_sort_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_auto_sort_mail_app"])))
    click_auto_sort_mail_app.click()
    time.sleep(2)
    Logging("3. Select Auto-Sort successfully")
    click_icon_add_auto_sort = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_plus_add_auto_sort"])))
    click_icon_add_auto_sort.click()
    time.sleep(2)
    Logging("4. Click Icon Add successfully")
    txt_input_form = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_from"])))
    txt_input_form.send_keys(data["send_mail"]["email_from"])
    Logging("5. Input Email From successfully")
    time.sleep(2)
    txt_input_to = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_to"])))
    txt_input_to.send_keys(data["send_mail"]["email_to"])
    Logging("6. Input Email To successfully")
    time.sleep(2)
    txt_input_subject = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["txt_subject_add_auto_sort"])))
    txt_input_subject.send_keys(data["send_mail"]["subject_auto_sort"])
    Logging("7. Input Subject Auto Sort successfully")
    time.sleep(2)
    click_btn_save_auto_sort = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["btn_save_auto_sort"])))
    click_btn_save_auto_sort.click()
    time.sleep(3)
    Logging("8. Save  Auto-Sort successfully")
    time.sleep(3)
    if 'QA Team Test' in driver.page_source :
        Logging("=> Add Auto-Sort => PASS")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["auto-sort"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Add Auto-Sort => FAIL")
        TestCase_LogResult(**data["testcase_result"]["mail_app"]["auto-sort"]["fail"])
    

    time.sleep(3) 
    Logging("------------------------------------------------------Delete Auto-Sort------------------------------------------------------")
    click_icon_edit_auto_sort = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_edit_auto_sort"])))
    click_icon_edit_auto_sort.click()
    time.sleep(1)
    Logging("1. Click Icon Edit Auto-Sort successfully")
    click_icon_delete_auto_sort = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_delete_auto_sort"])))
    click_icon_delete_auto_sort.click()
    time.sleep(3)
    Logging("2. Click Icon Delete Auto-Sort successfully")
    click_icon_done_delete_auto_sort = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["auto_sort_btn_done"])))
    click_icon_done_delete_auto_sort.click()
    time.sleep(1)
    Logging("3. Click Icon Done successfully")
    time.sleep(2)
    if 'QA Team Test' in driver.page_source :
        Logging("=> Delete Auto-Sort => FAIL")
    else:
        Logging("=> Delete Auto-Sort => PASS")
    time.sleep(1)
    driver.back()
    driver.back()
    time.sleep(3)
    

    
    Logging("------------------------------------------------------Delete Mail in Receipts------------------------------------------------------")
    time.sleep(1)
    click_icon_list_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["icon_list_mail"])))  
    click_icon_list_mail.click()
    time.sleep(2)
    driver.swipe(start_x=404, start_y=778, end_x=404, end_y=1400, duration=800)
    time.sleep(1)
    click_folder_inbox_mail_app = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_folder_receipts"])))
    click_folder_inbox_mail_app.click()
    time.sleep(1)
    Logging("1. Click Inbox successfully")
    time.sleep(1)
    select_mail3_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail1_delete"])))
    select_mail3_delete.click()
    select_mail_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail2_delete"])))
    select_mail_delete.click()
    select_mail2_delete = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_mail3_delete"])))
    select_mail2_delete.click()
    
    Logging("2. Select multiple email  successfully")
    click_icon_delete_mail= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["send_mail"]["select_icon_delete"])))
    click_icon_delete_mail.click()
    time.sleep(1)
    Logging("3. Click Icon Delete mail successfully")
    time.sleep(1)
    
   
    
    










#log_in_mail_app()
#send_mail_app()
#vacation_auto_replies_mail_app()
#auto_sort_mail_app()
